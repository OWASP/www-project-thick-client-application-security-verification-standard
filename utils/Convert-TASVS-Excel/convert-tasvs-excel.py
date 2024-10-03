import os
import shutil
import requests
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font


class ChecklistProcessor:
    """Class to process and download checklist data from GitHub markdown files."""

    def __init__(self, repo_document_root_url, tasvs_files):
        self.repo_document_root_url = repo_document_root_url
        self.tasvs_files = tasvs_files

    def get_testing_checklist(self, content):
        """Extract the 'Testing Checklist' section and its table from markdown content."""
        lines = content.split("\n")
        start_idx, end_idx = -1, -1
        checklist_lines = []

        # Locate the "Testing Checklist" section
        for i, line in enumerate(lines):
            if "## Testing Checklist" in line:
                start_idx = i
            elif start_idx != -1 and line.startswith("#"):
                end_idx = i
                break

        # Extract the content if we found a section
        if start_idx != -1:
            checklist_lines = (
                lines[start_idx + 1 : end_idx]
                if end_idx != -1
                else lines[start_idx + 1 :]
            )

        # Clean up the lines (remove empty lines or excess spaces)
        checklist_lines = [line.strip() for line in checklist_lines if line.strip()]

        # Parse the Markdown table from the checklist
        table_data = []
        in_table = False
        for line in checklist_lines:
            if "----" not in line and "TASVS-ID" not in line:  # Table row line
                row = [
                    cell.strip() for cell in line.split("|")[1:-1]
                ]  # Split and remove leading/trailing pipes
                table_data.append(row)
                in_table = True
            elif in_table and not line.startswith("|"):
                break  # Stop if we have exited the table section

        return table_data

    def process_files(self):
        """Download and process all markdown files, extracting checklist data."""
        checklist_data = []

        for file_name, sheet_name in self.tasvs_files:
            url = self.repo_document_root_url + file_name
            response = requests.get(url)

            if response.status_code == 200:
                content = response.text
                table_data = self.get_testing_checklist(content)
                if table_data:
                    checklist_data.append((table_data, sheet_name))
                    print(f"Checklist table extracted from {file_name}")
            else:
                print(
                    f"Failed to download {file_name} with status code {response.status_code}"
                )

        return checklist_data


class ExcelPopulator:
    """Class to handle the population of checklist data into an Excel template."""

    def __init__(self, template_path, output_file_path, testing_notes_map):
        self.template_path = template_path
        self.output_file_path = output_file_path
        self.testing_notes_map = testing_notes_map
        self._prepare_workbook()

    def _prepare_workbook(self):
        """Create a copy of the original Excel file to work on."""
        shutil.copy(self.template_path, self.output_file_path)

    def populate_spreadsheet(self, checklist_data, sheet_name):
        """Populate the extracted content into the corresponding Excel sheet."""
        wb = openpyxl.load_workbook(self.output_file_path)

        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return

        sheet = wb[sheet_name]
        dropdown_options = ["Failed", "N/A", "Pending", "Reviewed"]
        dropdown = DataValidation(
            type="list", formula1=f'"{",".join(dropdown_options)}"', allow_blank=True
        )
        dropdown.error = "Invalid entry, please select from the dropdown options."
        dropdown.prompt = "Please select an option from the list."
        sheet.add_data_validation(dropdown)

        notes_dict = {item[0]: item[1] for item in self.testing_notes_map}
        start_row = 12

        for row_data in checklist_data:
            sheet[f"B{start_row}"] = row_data[0] if row_data[0] else ""
            sheet[f"C{start_row}"] = row_data[1] if row_data[1] else ""
            sheet[f"D{start_row}"] = row_data[2] if row_data[2] else ""
            sheet[f"E{start_row}"] = row_data[3] if row_data[3] else ""
            sheet[f"F{start_row}"] = row_data[4] if row_data[4] else ""

            tasvs_id = row_data[0]
            if tasvs_id in notes_dict:
                sheet[f"K{start_row}"] = notes_dict[tasvs_id]

            non_empty_cells = sum(1 for item in row_data if item)

            if non_empty_cells == 2:
                for col in range(2, 3):  # Columns B (2) to C (3)
                    sheet.cell(row=start_row, column=col).font = Font(bold=True)
                sheet.row_dimensions[start_row].height = 26
            else:
                dropdown.add(sheet[f"G{start_row}"])
                sheet.row_dimensions[start_row].height = 90

            start_row += 1

        self._apply_global_formatting(sheet, start_row)
        wb.save(self.output_file_path)
        print(
            f"Data populated successfully into sheet '{sheet_name}' in {self.output_file_path}"
        )

    def _apply_global_formatting(self, sheet, end_row):
        """Apply global formatting across the sheet."""
        for row in range(12, end_row + 1):
            for col in range(2, 12):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )


class TASVSConversion:
    """Main class to manage the TASVS conversion process."""

    def __init__(
        self,
        repo_document_root_url,
        tasvs_files,
        template_path,
        output_file_path,
        testing_notes_map,
    ):
        self.checklist_processor = ChecklistProcessor(
            repo_document_root_url, tasvs_files
        )
        self.excel_populator = ExcelPopulator(
            template_path, output_file_path, testing_notes_map
        )

    def run(self):
        """Run the complete TASVS conversion process."""
        checklist_data = self.checklist_processor.process_files()

        for table_data, sheet_name in checklist_data:
            self.excel_populator.populate_spreadsheet(table_data, sheet_name)


if __name__ == "__main__":
    repo_document_root_url = "https://raw.githubusercontent.com/OWASP/www-project-thick-client-application-security-verification-standard/main/document/1.0/"
    tasvs_files = [
        ("04-TASVS-ARCH.md", "TASVS-ARCH"),
        ("05-TASVS-CODE.md", "TASVS-CODE"),
        ("06-TASVS-CONF.md", "TASVS-CONF"),
        ("07-TASVS-CRYPTO.md", "TASVS-CRYPTO"),
        ("08-TASVS-NETWORK.md", "TASVS-NETWORK"),
        ("09-TASVS-STORAGE.md", "TASVS-STORAGE"),
    ]
    template_path = os.path.join(
        os.path.dirname(os.path.realpath(__file__)), "TASVS_V0.99999999_orig.xlsx"
    )

    response = requests.get(
        "https://api.github.com/repos/OWASP/www-project-thick-client-application-security-verification-standard/releases/latest"
    )

    if response.status_code == 200:
        # Parse the JSON response
        latest_release = response.json()
        latest_tag = latest_release["tag_name"]
    else:
        print(f"Failed to retrieve latest release. Status code: {response.status_code}")
        # default the tag to something so it looks good in the filename
        latest_tag = "v1.0"

    output_file_path = os.path.join(
        os.path.dirname(os.path.realpath(__file__)), f"TASVS_{latest_tag}.xlsx"
    )

    # map for data to be inserted into col K "Testing notes". Format:
    # (TASVS-ID, Note, [hyperlink])
    #
    # fmt is to tell black to ignore the formatting
    # fmt: off
    testing_notes_map = [
    ("TASVS-ARCH-1.1", "TASVS-ARCH-1.1 satisfied by TASVS-ARCH-1.3", ""),
    ("TASVS-ARCH-1.2", "Fidelity score >= 80%", ""),
    ("TASVS-ARCH-1.3", "External dependencies can be defined as out-of-scope where appropriate.", ""),
    ("TASVS-ARCH-1.6", "Within prior 90 days", ""),
    ("TASVS-CODE-1.1", "Recommended: OWASP ASVS", "https://owasp.org/www-project-application-security-verification-standard/"),
    ("TASVS-CODE-2.5", "Example: C Hardening Cheat Sheet", "https://cheatsheetseries.owasp.org/cheatsheets/C-Based_Toolchain_Hardening_Cheat_Sheet.html"),
    ("TASVS-CODE-3.1", "Recommended: OWASP Dependency-Check", "https://github.com/jeremylong/DependencyCheck"),
    ("TASVS-CODE-3.3", "Recommended: BinSkim", "https://github.com/microsoft/binskim"),
    ("TASVS-CODE-3.4", "E.g: Python=Bandit, C#=Security Code Scan. Fallback to OWASP cheatsheets and/or use SemGrep.", ""),
    ("TASVS-CODE-3.6", "Case Study", "https://www.henricodolfing.com/2019/06/project-failure-case-study-knight-capital.html"),
    ("TASVS-CODE-4.10", "Example: batbadbut research", "https://flatt.tech/research/posts/batbadbut-you-cant-securely-execute-commands-on-windows/"),
    ("TASVS-CODE-6.2", "Also satisfies TASVS-CODE-6.1", ""),
    ]
    # fmt: on

    conversion = TASVSConversion(
        repo_document_root_url,
        tasvs_files,
        template_path,
        output_file_path,
        testing_notes_map,
    )
    conversion.run()
